"""Convert office documents (Word, PPT, Excel, OpenDocument) to PDF for OCR.

Uses LibreOffice in headless mode. The resulting PDF can be passed to the
existing PDF/image pipeline (PageLoader, MaaS).

Supported extensions: .doc, .docx, .pptx, .xls, .xlsx, .ods
(PDF and images are supported natively.)

Requires: LibreOffice installed. On Windows, if soffice is not on PATH,
common install paths (e.g. Program Files\\LibreOffice\\program\\soffice.exe)
are tried automatically.
"""

from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import List, Optional, Tuple

from glmocr.utils.logging import get_logger

logger = get_logger(__name__)

# Extensions that should be converted to PDF before loading (Word, PPT, Excel)
SUPPORTED_DOC_EXTENSIONS = {".doc", ".docx", ".pptx", ".xls", ".xlsx", ".ods"}
ALL_DOC_EXTENSIONS = SUPPORTED_DOC_EXTENSIONS

EXCEL_EXTENSIONS = {".xls", ".xlsx", ".ods"}


def _convert_to_xlsx(file_path: str, temp_dir: str, timeout_seconds: int = 60) -> str:
    """Convert .xls / .ods to .xlsx using LibreOffice so openpyxl can process it.

    Args:
        file_path: Path to .xls or .ods file.
        temp_dir: Temporary directory for output.
        timeout_seconds: LibreOffice timeout.

    Returns:
        Path to the converted .xlsx file.
    """
    cmd = _get_libreoffice_cmd()
    args = [
        cmd,
        "--headless",
        "--norestore",
        "--nofirststartwizard",
        "--nologo",
        "--convert-to",
        "xlsx",
        "--outdir",
        temp_dir,
        str(Path(file_path).resolve()),
    ]
    logger.debug("Converting to xlsx: %s", " ".join(args))
    subprocess.run(args, capture_output=True, text=True, timeout=timeout_seconds, check=True)
    xlsx_name = Path(file_path).stem + ".xlsx"
    xlsx_path = os.path.join(temp_dir, xlsx_name)
    if not os.path.isfile(xlsx_path):
        candidates = [f for f in os.listdir(temp_dir) if f.lower().endswith(".xlsx")]
        if candidates:
            xlsx_path = os.path.join(temp_dir, candidates[0])
        else:
            raise RuntimeError(f"LibreOffice did not produce an xlsx for: {file_path}")
    return xlsx_path


def _prepare_excel_fit_to_page(file_path: str, temp_dir: str) -> str:
    """Apply 'fit to 1 page wide' print scaling on every sheet via openpyxl.

    This mirrors Excel's print-preview "Fit Sheet on One Page Wide" so that
    LibreOffice will produce a properly scaled PDF instead of cutting off
    wide columns.

    .xls and .ods files are first converted to .xlsx via LibreOffice so that
    openpyxl can apply the page setup.

    Args:
        file_path: Path to the original Excel file (.xlsx / .xls / .ods).
        temp_dir: Temporary directory to write intermediate/modified files.

    Returns:
        Path to the .xlsx file with fit-to-page scaling applied.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed; skipping fit-to-page scaling for %s", file_path)
        return file_path

    path = Path(file_path)
    ext = path.suffix.lower()

    # Convert .xls / .ods → .xlsx first so openpyxl can open it
    if ext in (".xls", ".ods"):
        try:
            file_path = _convert_to_xlsx(file_path, temp_dir)
        except Exception as e:
            logger.warning("xlsx conversion failed for %s: %s; using original", file_path, e)
            return file_path

    try:
        wb = load_workbook(file_path)
        for ws in wb.worksheets:
            if ws is None:
                continue
            try:
                max_col = ws.max_column or 1

                # Wide sheets: landscape gives more horizontal print area
                if max_col > 8:
                    ws.page_setup.orientation = "landscape"

                # "Fit to 1 page wide, auto height": LibreOffice scales based on
                # actual column widths (more accurate than a column-count estimate).
                # fitToHeight=0 means unlimited pages tall — no row compression.
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_setup.fitToWidth = 1
                ws.page_setup.fitToHeight = 0
                logger.debug(
                    "Sheet '%s': fitToWidth=1 fitToHeight=0 landscape=%s",
                    getattr(ws, 'title', '?'), max_col > 8,
                )
            except Exception as ws_err:
                logger.debug("Skipping sheet %s: %s", getattr(ws, 'title', '?'), ws_err)
        out_path = os.path.join(temp_dir, Path(file_path).stem + "_scaled.xlsx")
        wb.save(out_path)
        logger.debug("Saved fit-to-page Excel to: %s", out_path)
        return out_path
    except Exception as e:
        logger.warning("fit-to-page scaling failed for %s: %s; using original", file_path, e)
        return file_path

# LibreOffice command (soffice on some systems, libreoffice on others)
_LIBREOFFICE_CMD: Optional[str] = None


def _get_windows_libreoffice_paths() -> List[str]:
    """Return candidate paths for soffice.exe on Windows (not on PATH)."""
    if sys.platform != "win32":
        return []
    candidates = []
    for base in (
        os.environ.get("ProgramFiles", "C:\\Program Files"),
        os.environ.get("ProgramFiles(x86)", "C:\\Program Files (x86)"),
    ):
        if not os.path.isdir(base):
            continue
        try:
            for name in os.listdir(base):
                if name.lower().startswith("libreoffice"):
                    exe_path = os.path.join(base, name, "program", "soffice.exe")
                    if os.path.isfile(exe_path):
                        candidates.append(exe_path)
        except OSError:
            continue
    # Prefer Program Files over Program Files (x86)
    candidates.sort(key=lambda x: ("x86" in x.lower(), len(x)))
    return candidates


def _get_libreoffice_cmd() -> str:
    global _LIBREOFFICE_CMD
    if _LIBREOFFICE_CMD is not None:
        return _LIBREOFFICE_CMD
    # 1) Try PATH
    for cmd in ("soffice", "libreoffice"):
        try:
            subprocess.run(
                [cmd, "--version"],
                capture_output=True,
                timeout=5,
                check=True,
            )
            _LIBREOFFICE_CMD = cmd
            return cmd
        except (FileNotFoundError, subprocess.CalledProcessError, subprocess.TimeoutExpired):
            continue
    # 2) On Windows, try common install paths
    for exe_path in _get_windows_libreoffice_paths():
        try:
            subprocess.run(
                [exe_path, "--version"],
                capture_output=True,
                timeout=5,
                check=True,
            )
            _LIBREOFFICE_CMD = exe_path
            logger.info("Using LibreOffice at: %s", exe_path)
            return exe_path
        except (OSError, subprocess.CalledProcessError, subprocess.TimeoutExpired):
            continue
    raise RuntimeError(
        "LibreOffice is required for Word/PPT/Excel/OpenDocument (doc/docx/pptx/xls/xlsx/ods) conversion. "
        "Please install LibreOffice from https://www.libreoffice.org/ "
        "On Windows, install to default location or add its 'program' folder to PATH."
    )


def is_document_path(path: str) -> bool:
    """Return True if the path has an extension that needs conversion to PDF."""
    ext = Path(path).suffix.lower()
    return ext in ALL_DOC_EXTENSIONS


def convert_to_pdf(
    file_path: str,
    timeout_seconds: int = 120,
) -> Tuple[str, str]:
    """Convert a document (doc, docx, pptx, xls, xlsx, ods) to PDF using LibreOffice.

    Args:
        file_path: Path to the document.
        timeout_seconds: Max time for conversion.

    Returns:
        (pdf_path, temp_dir). The PDF is created inside temp_dir. Caller should
        remove temp_dir when done (e.g. shutil.rmtree(temp_dir, ignore_errors=True)).

    Raises:
        FileNotFoundError: If file_path does not exist.
        ValueError: If file extension is not supported.
        RuntimeError: If LibreOffice is not available or conversion fails.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    ext = path.suffix.lower()
    if ext not in ALL_DOC_EXTENSIONS:
        raise ValueError(
            f"Unsupported document extension: {ext}. "
            f"Supported: {sorted(ALL_DOC_EXTENSIONS)}"
        )

    cmd = _get_libreoffice_cmd()
    temp_dir = tempfile.mkdtemp(prefix="glmocr_doc_")
    try:
        # For Excel files, apply fit-to-page scaling before conversion so that
        # wide sheets render like print preview (1 page wide, N pages tall).
        if ext in EXCEL_EXTENSIONS:
            file_path = _prepare_excel_fit_to_page(file_path, temp_dir)

        # For spreadsheets, use calc_pdf_Export filter to ensure ALL sheets are
        # exported (default "pdf" target may skip non-active sheets in some
        # LibreOffice versions).
        convert_to = (
            "pdf:calc_pdf_Export" if ext in EXCEL_EXTENSIONS else "pdf"
        )

        args = [
            cmd,
            "--headless",
            "--norestore",
            "--nofirststartwizard",
            "--nologo",
            "--convert-to",
            convert_to,
            "--outdir",
            temp_dir,
            str(Path(file_path).resolve()),  # use the (possibly scaled) file
        ]
        logger.debug("Running: %s", " ".join(args))
        subprocess.run(
            args,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            check=True,
        )
        # Output filename is same stem + .pdf (based on the converted file)
        pdf_name = Path(file_path).stem + ".pdf"
        pdf_path = os.path.join(temp_dir, pdf_name)
        if not os.path.isfile(pdf_path):
            candidates = [f for f in os.listdir(temp_dir) if f.lower().endswith(".pdf")]
            if candidates:
                pdf_path = os.path.join(temp_dir, candidates[0])
            else:
                raise RuntimeError(
                    f"LibreOffice did not produce a PDF for: {file_path}"
                )
        return pdf_path, temp_dir
    except subprocess.TimeoutExpired:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise RuntimeError(
            f"Document conversion timed out after {timeout_seconds}s: {file_path}"
        ) from None
    except subprocess.CalledProcessError as e:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise RuntimeError(
            f"LibreOffice conversion failed for {file_path}: {e.stderr or e}"
        ) from e
